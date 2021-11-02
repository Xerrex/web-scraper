# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
import openpyxl



class VehiclesPipeline:

    def open_spider(self, spider):
        self.vehicles_wb = openpyxl.Workbook()
    
    def close_spider(self, spider):
        self.vehicles_wb.save("data/vehicles.xlsx")
    
    def process_item(self, item, spider):
        active_sheet = self.vehicles_wb.active
        
        active_sheet.cell(1,1).value = "Name"
        active_sheet.cell(1,2).value = "Price"
        active_sheet.cell(1,3).value = "Vin Number"
        active_sheet.cell(1,4).value = "Vehicle Summary"
        active_sheet.cell(1,5).value = "Top Features & Specs"
        
        vehicle_item_adapter = ItemAdapter(item)
        price = vehicle_item_adapter.get('price')
        name = vehicle_item_adapter.get('name')
        vin_number = vehicle_item_adapter.get('vin_number')

        vehicle_summary = self.process_vehicle_summary(vehicle_item_adapter.get('vehicle_summary'))
        top_features_specs = self.process_features_and_Specs(vehicle_item_adapter.get('top_features_specs'))
        
        active_sheet.append([name, price, vin_number, vehicle_summary, top_features_specs])
        
        return item

    def process_features_and_Specs(self, top_features_specs):
        """Make the top_features_specs into a sentence
        """
        features_Specs = []
        for k,v in top_features_specs.items():
            ft_spec = f"{k}: {','.join(v)}"
            features_Specs.append(ft_spec)
        return ".".join(features_Specs)
    
    def process_vehicle_summary(self, vehicle_summary):
        """Transform the Vehicle summary into a sentence
        """
        return ",".join(vehicle_summary)
    